USE DATABASE EXPERIMENT_TEAM_DB;
USE SCHEMA PUBLIC;

CREATE OR REPLACE TABLE OPEN_MARKET_CONSTRUCTION_CODE_EXTRACTION (
    ACCOUNT_FOLDER_NAME           STRING,
    CONSTRUCTION_CODE             VARIANT,
    RAW_CONSTRUCTION_TYPES        VARIANT,
    RAW_TO_NORMALIZED_MAPPING     VARIANT,
    EVIDENCE                      VARIANT,
    SOURCE_FILES                  VARIANT,
    FILES_FOUND                   NUMBER,
    FILES_PROCESSED               NUMBER,
    FILES_SKIPPED                 VARIANT,
    CONFIDENCE                    FLOAT,
    PROCESSED_AT                  TIMESTAMP_NTZ,
    ERROR_MESSAGE                 STRING
);

CREATE OR REPLACE PROCEDURE RUN_OPEN_MARKET_CONSTRUCTION_CODE_EXTRACTION(
    STAGE_ROOT STRING,
    MAX_FOLDERS NUMBER
)
RETURNS STRING
LANGUAGE PYTHON
RUNTIME_VERSION = '3.11'
PACKAGES = ('snowflake-snowpark-python', 'openpyxl')
HANDLER = 'run'
EXECUTE AS OWNER
AS
$$
import json
import re
import io
from collections import Counter, defaultdict
from datetime import datetime

from openpyxl import load_workbook
from snowflake.snowpark.files import SnowflakeFile


# ------------------------------------------------------------
# 1. STANDARD NORMALIZATION MAP
# ------------------------------------------------------------
# Exact raw values and common abbreviations/synonyms that we already know.
# You can keep extending this over time as new values appear.
EXACT_NORMALIZATION_MAP = {
    "UNKNOWN": "Unknown",

    "FRAME": "Frame",
    "WOOD FRAME": "Frame",
    "5 STORY WRAP": "Frame",

    "JM": "Joisted Masonry",
    "JOISTED MASONRY": "Joisted Masonry",

    "NON COMBUSTIBLE": "Non-Combustible",
    "NON-COMBUSTIBLE": "Non-Combustible",
    "NC": "Non-Combustible",
    "PEMB": "Non-Combustible",
    "METAL": "Non-Combustible",

    "MNC": "Masonry Non-Combustible",
    "MASONRY NON COMBUSTIBLE": "Masonry Non-Combustible",
    "MASONRY NON-COMBUSTIBLE": "Masonry Non-Combustible",
    "MASONRY": "Masonry Non-Combustible",
    "BRICK WITH MASONRY": "Masonry Non-Combustible",
    "BRICK/BLOCK": "Masonry Non-Combustible",
    "BRICK ON BLOCK": "Masonry Non-Combustible",
    "CMU": "Masonry Non-Combustible",
    "CMU BLOCK": "Masonry Non-Combustible",
    "CMU/BRICK ON BLOCK": "Masonry Non-Combustible",
    "CMU BLOCK WITH STUCCO FACADE": "Masonry Non-Combustible",
    "CMU BLOCK W/ STUCCO FACADE": "Masonry Non-Combustible",
    "CONCRETE BLOCK STUCCO": "Masonry Non-Combustible",
    "MASONRY WITH STUCCO": "Masonry Non-Combustible",
    "BRICK": "Masonry Non-Combustible",

    "CONCRETE": "Modified Fire Resistive",
    "REINFORCED CONCRETE/PODIUM": "Modified Fire Resistive",
    "CONCRETE PODIUM WITH BRICK": "Modified Fire Resistive",
    "STEEL FRAME/CONCRETE": "Modified Fire Resistive",

    "MODIFIED FIRE RESISTIVE": "Modified Fire Resistive",
    "FIRE RESISTIVE": "Fire Resistive",
    "HEAVY TIMBER JOISTED MASONRY": "Heavy Timber Joisted Masonry",
    "SUPERIOR NON-COMBUSTIBLE": "Superior Non-Combustible",
    "SUPERIOR NON COMBUSTIBLE": "Superior Non-Combustible",
    "SUPERIOR MASONRY NON-COMBUSTIBLE": "Superior Masonry Non-Combustible",
    "SUPERIOR MASONRY NON COMBUSTIBLE": "Superior Masonry Non-Combustible",
}


# ------------------------------------------------------------
# 2. HELPER FUNCTIONS
# ------------------------------------------------------------
def clean_text(value):
    if value is None:
        return None
    value = str(value).strip()
    if not value:
        return None
    value = re.sub(r'\s+', ' ', value)
    return value


def canonical_key(value):
    """
    Used only for matching.
    Keeps the display/original value separate.
    """
    value = clean_text(value)
    if not value:
        return None

    value = value.upper()
    value = value.replace("–", "-").replace("—", "-")
    value = value.replace("&", " AND ")
    value = value.replace("FAÇADE", "FACADE")
    value = re.sub(r'\s+', ' ', value).strip()

    return value


def normalize_construction_type(raw_value):
    """
    Returns normalized construction code based on:
    1. exact mappings
    2. rule-based fallback for unseen but obvious descriptions
    """
    key = canonical_key(raw_value)

    if not key:
        return "Unknown"

    if key in EXACT_NORMALIZATION_MAP:
        return EXACT_NORMALIZATION_MAP[key]

    # Rule-based fallbacks for unseen values
    if "WOOD" in key and "FRAME" in key:
        return "Frame"

    if key == "FRAME":
        return "Frame"

    if "JOISTED MASONRY" in key:
        return "Joisted Masonry"

    if "HEAVY TIMBER" in key:
        return "Heavy Timber Joisted Masonry"

    if "FIRE RESISTIVE" in key and "MODIFIED" in key:
        return "Modified Fire Resistive"

    if "FIRE RESISTIVE" in key:
        return "Fire Resistive"

    if "SUPERIOR" in key and "MASONRY" in key and "NON" in key and "COMBUST" in key:
        return "Superior Masonry Non-Combustible"

    if "SUPERIOR" in key and "NON" in key and "COMBUST" in key:
        return "Superior Non-Combustible"

    if "NON" in key and "COMBUST" in key and "MASONRY" in key:
        return "Masonry Non-Combustible"

    if "NON" in key and "COMBUST" in key:
        return "Non-Combustible"

    if any(token in key for token in [
        "CMU", "MASONRY", "BRICK", "BLOCK", "TILT WALL", "TILT-WALL",
        "TILT UP", "TILT-UP", "CONCRETE TILT", "CB-TILT", "CB TILT"
    ]):
        return "Masonry Non-Combustible"

    if "CONCRETE" in key or "STEEL FRAME" in key:
        return "Modified Fire Resistive"

    if "METAL" in key:
        return "Non-Combustible"

    return "Unknown"


def likely_construction_header(value):
    """
    Detect likely construction-related Excel column headers.
    """
    if value is None:
        return False

    h = str(value).strip().lower()

    header_keywords = [
        "construction",
        "construction description",
        "construction type",
        "construction code",
        "iso construction",
        "const.",
        "const ",
        "const_",
        "building construction"
    ]

    return any(k in h for k in header_keywords)


def add_raw_value(
    raw_value,
    raw_counter,
    normalized_counter,
    raw_to_normalized_mapping,
    evidence_rows,
    source_file,
    evidence_text
):
    raw_value = clean_text(raw_value)

    if not raw_value:
        return

    # Ignore obvious blanks / non-values
    invalid_values = {"N/A", "NA", "NULL", "NONE", "-", "--"}
    if raw_value.upper() in invalid_values:
        return

    normalized = normalize_construction_type(raw_value)

    raw_counter[raw_value] += 1
    normalized_counter[normalized] += 1
    raw_to_normalized_mapping[raw_value] = normalized

    evidence_rows.append({
        "source_file": source_file,
        "raw_value": raw_value,
        "normalized_code": normalized,
        "evidence": evidence_text
    })


def escape_sql_string(value):
    if value is None:
        return ""
    return str(value).replace("'", "''")


def split_stage_root(stage_root):
    """
    Example:
      @OPEN_MARKET_SUBMISSION/Open_Market
    Returns:
      stage_name = @OPEN_MARKET_SUBMISSION
      relative_root = Open_Market
    """
    stage_root = stage_root.rstrip("/")

    if "/" not in stage_root:
        return stage_root, ""

    first_slash = stage_root.find("/")
    stage_name = stage_root[:first_slash]
    relative_root = stage_root[first_slash + 1:].strip("/")

    return stage_name, relative_root


def get_relative_path_from_list_name(list_name, stage_name):
    """
    LIST output may contain paths like:
      open_market_submission/Open_Market/Folder/file.xlsx
    We convert it to:
      Open_Market/Folder/file.xlsx
    """
    name = list_name.replace("\\", "/")

    stage_without_at = stage_name.replace("@", "")
    stage_without_at = stage_without_at.split(".")[-1]

    possible_prefixes = [
        stage_without_at + "/",
        stage_without_at.lower() + "/",
        stage_without_at.upper() + "/"
    ]

    for prefix in possible_prefixes:
        if name.startswith(prefix):
            return name[len(prefix):]

    return name


def get_folder_name(relative_path, relative_root):
    """
    relative_path:
      Open_Market/A162361_KHP_Capital_Partners/file.xlsx

    relative_root:
      Open_Market

    returns:
      A162361_KHP_Capital_Partners
    """
    path = relative_path.replace("\\", "/").strip("/")

    if relative_root:
        root_prefix = relative_root.strip("/") + "/"
        if path.startswith(root_prefix):
            path = path[len(root_prefix):]

    parts = path.split("/")

    if len(parts) < 2:
        return None

    return parts[0]


def get_file_name(relative_path):
    return relative_path.replace("\\", "/").split("/")[-1]


def get_extension(file_name):
    file_name = file_name.lower()
    if "." not in file_name:
        return ""
    return "." + file_name.split(".")[-1]


# ------------------------------------------------------------
# 3. XLSX HANDLING
# ------------------------------------------------------------
def extract_from_xlsx(scoped_file_url, relative_path):
    """
    Reads XLSX using openpyxl.
    Counts every non-empty value under construction-like columns.
    """
    found_values = []

    with SnowflakeFile.open(scoped_file_url, "rb") as f:
        data = f.read()

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)

        header_row = None
        header_idx_to_name = {}

        # Search first 30 rows for a likely header row
        for row_number, row in enumerate(rows, start=1):
            if row_number > 30:
                break

            current_header_map = {}
            for idx, cell_value in enumerate(row):
                if likely_construction_header(cell_value):
                    current_header_map[idx] = clean_text(cell_value)

            if current_header_map:
                header_row = row_number
                header_idx_to_name = current_header_map
                break

        if header_row is None:
            continue

        # Re-open worksheet iterator because previous rows were consumed
        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_number <= header_row:
                continue

            for col_idx, header_name in header_idx_to_name.items():
                if col_idx < len(row):
                    value = clean_text(row[col_idx])

                    if value:
                        found_values.append({
                            "raw_value": value,
                            "sheet_name": ws.title,
                            "row_number": row_number,
                            "header_name": header_name,
                            "evidence": f"Sheet '{ws.title}', row {row_number}, column '{header_name}' = '{value}'"
                        })

    return found_values


# ------------------------------------------------------------
# 4. AI_EXTRACT HANDLING FOR PDF / DOC / DOCX / EML
# ------------------------------------------------------------
def ai_extract_from_file(session, stage_name, relative_path):
    """
    Asks AI_EXTRACT for raw construction values and supporting evidence.
    """
    relative_path_sql = escape_sql_string(relative_path)
    stage_name_sql = escape_sql_string(stage_name)

    response_format = """
    {
      'schema': {
        'type': 'object',
        'properties': {
          'construction_values': {
            'description': 'List every construction-related value appearing in the document. Look for fields or columns such as Construction, Construction Description, Construction Type, Construction Code, ISO Construction, Building Construction, or close synonyms. Return the raw values exactly as written in the document. Include abbreviations such as MNC, JM, PEMB, CB-TILT, etc. Do not invent values.',
            'type': 'array'
          },
          'evidence': {
            'description': 'Return a short evidence snippet or nearby text that supports the construction value extraction.',
            'type': 'array'
          }
        }
      }
    }
    """

    sql = f"""
        SELECT AI_EXTRACT(
            file => TO_FILE('{stage_name_sql}', '{relative_path_sql}'),
            responseFormat => {response_format},
            scores => TRUE
        ) AS R
    """

    result = session.sql(sql).collect()[0]["R"]

    if isinstance(result, str):
        result = json.loads(result)

    response = result.get("response", {}) if result else {}
    scoring = result.get("scoring", {}) if result else {}
    error = result.get("error") if result else None

    values = response.get("construction_values", []) or []
    evidence = response.get("evidence", []) or []

    score = None
    try:
        score = scoring.get("scores", {}).get("construction_values", {}).get("score")
    except Exception:
        score = None

    return values, evidence, score, error


# ------------------------------------------------------------
# 5. AI_PARSE_DOCUMENT FALLBACK FOR PDF / DOCX
# ------------------------------------------------------------
def ai_parse_document_text(session, stage_name, relative_path):
    """
    Uses AI_PARSE_DOCUMENT only for file types it supports.
    """
    relative_path_sql = escape_sql_string(relative_path)
    stage_name_sql = escape_sql_string(stage_name)

    sql = f"""
        SELECT AI_PARSE_DOCUMENT(
            TO_FILE('{stage_name_sql}', '{relative_path_sql}'),
            {{'mode': 'LAYOUT'}}
        ) AS R
    """

    result = session.sql(sql).collect()[0]["R"]

    if isinstance(result, str):
        result = json.loads(result)

    # Newer output can contain either content or pages
    if not result:
        return ""

    if "content" in result and result["content"]:
        return str(result["content"])

    pages = result.get("pages", [])
    combined = []

    for p in pages:
        content = p.get("content")
        if content:
            combined.append(str(content))

    return "\n".join(combined)


def extract_known_values_from_text(text):
    """
    Fallback scan of parsed text for known values.
    This is intentionally conservative.
    """
    if not text:
        return []

    extracted = []

    known_raw_terms = [
        "MNC",
        "JM",
        "PEMB",
        "Frame",
        "Wood Frame",
        "Non Combustible",
        "Non-Combustible",
        "Masonry",
        "Concrete tilt-up",
        "Concrete Tilt-Up",
        "Concrete Tilt Wall",
        "CB-TILT",
        "CB Tilt Wall",
        "CMU",
        "CMU Block",
        "Brick/Block",
        "Brick with Masonry",
        "Brick on Block",
        "Tilt Wall",
        "Tilt Wall Concrete",
        "Steel Frame/Concrete"
    ]

    upper_text = text.upper()

    for term in known_raw_terms:
        count = upper_text.count(term.upper())
        if count > 0:
            for _ in range(count):
                extracted.append(term)

    return extracted


# ------------------------------------------------------------
# 6. MAIN PROCEDURE
# ------------------------------------------------------------
def run(session, STAGE_ROOT, MAX_FOLDERS):

    stage_name, relative_root = split_stage_root(STAGE_ROOT)

    # Step A: list all files recursively under the root
    list_sql = f"LIST {STAGE_ROOT}"
    list_rows = session.sql(list_sql).collect()

    all_files = []

    for row in list_rows:
        # LIST output first column is usually name
        list_name = row[0]
        relative_path = get_relative_path_from_list_name(list_name, stage_name)

        folder_name = get_folder_name(relative_path, relative_root)

        if not folder_name:
            continue

        file_name = get_file_name(relative_path)
        extension = get_extension(file_name)

        all_files.append({
            "folder_name": folder_name,
            "relative_path": relative_path,
            "file_name": file_name,
            "extension": extension
        })

    files_by_folder = defaultdict(list)

    for f in all_files:
        files_by_folder[f["folder_name"]].append(f)

    folder_names = sorted(files_by_folder.keys())

    if MAX_FOLDERS is not None and int(MAX_FOLDERS) > 0:
        folder_names = folder_names[:int(MAX_FOLDERS)]

    processed_folder_count = 0
    total_files_processed = 0

    supported_ai_extract_extensions = {".pdf", ".doc", ".docx", ".eml"}
    parse_supported_extensions = {".pdf", ".docx"}
    xlsx_extensions = {".xlsx"}

    for folder_name in folder_names:

        folder_files = files_by_folder[folder_name]

        raw_counter = Counter()
        normalized_counter = Counter()
        raw_to_normalized_mapping = {}
        evidence_rows = []
        source_files = set()
        skipped_files = []
        confidence_scores = []
        error_messages = []

        files_found = len(folder_files)
        files_processed = 0

        for file_info in folder_files:
            relative_path = file_info["relative_path"]
            file_name = file_info["file_name"]
            extension = file_info["extension"]

            try:
                # ------------------------------------------------
                # XLSX
                # ------------------------------------------------
                if extension in xlsx_extensions:
                    scoped_url_sql = f"""
                        SELECT BUILD_SCOPED_FILE_URL(
                            {stage_name},
                            '{escape_sql_string(relative_path)}'
                        ) AS FILE_URL
                    """
                    scoped_file_url = session.sql(scoped_url_sql).collect()[0]["FILE_URL"]

                    xlsx_values = extract_from_xlsx(scoped_file_url, relative_path)

                    for item in xlsx_values:
                        add_raw_value(
                            raw_value=item["raw_value"],
                            raw_counter=raw_counter,
                            normalized_counter=normalized_counter,
                            raw_to_normalized_mapping=raw_to_normalized_mapping,
                            evidence_rows=evidence_rows,
                            source_file=file_name,
                            evidence_text=item["evidence"]
                        )

                    files_processed += 1
                    source_files.add(file_name)

                # ------------------------------------------------
                # PDF / DOC / DOCX / EML via AI_EXTRACT
                # ------------------------------------------------
                elif extension in supported_ai_extract_extensions:
                    values, evidence, score, ai_error = ai_extract_from_file(
                        session=session,
                        stage_name=stage_name,
                        relative_path=relative_path
                    )

                    if ai_error:
                        error_messages.append(f"{file_name}: AI_EXTRACT error: {ai_error}")

                    # Add AI extracted values, counted once per occurrence returned
                    for idx, value in enumerate(values):
                        evidence_text = None

                        if idx < len(evidence):
                            evidence_text = evidence[idx]
                        elif evidence:
                            evidence_text = evidence[0]
                        else:
                            evidence_text = f"AI_EXTRACT identified '{value}' in {file_name}"

                        add_raw_value(
                            raw_value=value,
                            raw_counter=raw_counter,
                            normalized_counter=normalized_counter,
                            raw_to_normalized_mapping=raw_to_normalized_mapping,
                            evidence_rows=evidence_rows,
                            source_file=file_name,
                            evidence_text=evidence_text
                        )

                    if score is not None:
                        confidence_scores.append(float(score))

                    # Optional fallback: parse PDF/DOCX and scan known construction terms
                    if extension in parse_supported_extensions:
                        parsed_text = ai_parse_document_text(
                            session=session,
                            stage_name=stage_name,
                            relative_path=relative_path
                        )

                        fallback_values = extract_known_values_from_text(parsed_text)

                        for value in fallback_values:
                            add_raw_value(
                                raw_value=value,
                                raw_counter=raw_counter,
                                normalized_counter=normalized_counter,
                                raw_to_normalized_mapping=raw_to_normalized_mapping,
                                evidence_rows=evidence_rows,
                                source_file=file_name,
                                evidence_text=f"AI_PARSE_DOCUMENT fallback text scan found '{value}' in {file_name}"
                            )

                    files_processed += 1
                    source_files.add(file_name)

                # ------------------------------------------------
                # Unsupported extensions
                # ------------------------------------------------
                else:
                    skipped_files.append({
                        "file_name": file_name,
                        "reason": f"Unsupported extension: {extension if extension else 'no extension'}"
                    })

            except Exception as e:
                skipped_files.append({
                    "file_name": file_name,
                    "reason": str(e)
                })
                error_messages.append(f"{file_name}: {str(e)}")

        # Average confidence only from AI_EXTRACT outputs
        avg_confidence = None
        if confidence_scores:
            avg_confidence = sum(confidence_scores) / len(confidence_scores)

        # Avoid duplicate evidence rows explosion:
        # keep all evidence, but still as JSON array for auditability
        insert_sql = f"""
            INSERT INTO OPEN_MARKET_CONSTRUCTION_CODE_EXTRACTION (
                ACCOUNT_FOLDER_NAME,
                CONSTRUCTION_CODE,
                RAW_CONSTRUCTION_TYPES,
                RAW_TO_NORMALIZED_MAPPING,
                EVIDENCE,
                SOURCE_FILES,
                FILES_FOUND,
                FILES_PROCESSED,
                FILES_SKIPPED,
                CONFIDENCE,
                PROCESSED_AT,
                ERROR_MESSAGE
            )
            SELECT
                '{escape_sql_string(folder_name)}',
                PARSE_JSON('{escape_sql_string(json.dumps(dict(normalized_counter)))}'),
                PARSE_JSON('{escape_sql_string(json.dumps(dict(raw_counter)))}'),
                PARSE_JSON('{escape_sql_string(json.dumps(raw_to_normalized_mapping))}'),
                PARSE_JSON('{escape_sql_string(json.dumps(evidence_rows))}'),
                PARSE_JSON('{escape_sql_string(json.dumps(sorted(list(source_files))))}'),
                {files_found},
                {files_processed},
                PARSE_JSON('{escape_sql_string(json.dumps(skipped_files))}'),
                {"NULL" if avg_confidence is None else avg_confidence},
                CURRENT_TIMESTAMP(),
                {"NULL" if not error_messages else "'" + escape_sql_string(" | ".join(error_messages)) + "'"}
        """

        session.sql(insert_sql).collect()

        processed_folder_count += 1
        total_files_processed += files_processed

    return (
        f"Completed construction code extraction. "
        f"Folders processed: {processed_folder_count}. "
        f"Files processed: {total_files_processed}."
    )
$$;



CALL RUN_OPEN_MARKET_CONSTRUCTION_CODE_EXTRACTION(
    '@OPEN_MARKET_SUBMISSION/Open_Market',
    10
);
